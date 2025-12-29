import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class ClientExporter:

    @staticmethod
    def export_to_excel(clients, filepath):
        """Exporter la liste des clients en Excel"""
        try:
            # Créer un classeur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clients"

            # En-têtes
            headers = ["ID", "Nom", "Prénom", "Téléphone", "Email", "Adresse", "Ville", "Code Postal", "Date Création"]
            ws.append(headers)

            # Formater les en-têtes
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajouter les données
            for client in clients:
                ws.append([
                    client.get('id', ''),
                    client.get('nom', ''),
                    client.get('prenom', ''),
                    client.get('telephone', ''),
                    client.get('email', ''),
                    client.get('adresse', ''),
                    client.get('ville', ''),
                    client.get('code_postal', ''),
                    client.get('created_at', '')
                ])

            # Ajuster les largeurs de colonnes
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 18

            # Centrer les colonnes ID et Code Postal
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")

            wb.save(filepath)
            return True, f"Export réussi : {filepath}"
        except Exception as e:
            return False, f"Erreur d'export : {str(e)}"

    @staticmethod
    def export_client_history(client_name, history, filepath):
        """Exporter l'historique des achats d'un client en Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historique"

            # Titre
            ws['A1'] = f"Historique des achats - {client_name}"
            ws['A1'].font = Font(bold=True, size=12)

            # En-têtes
            headers = ["N° Facture", "Date", "Montant", "Statut"]
            ws.append([])
            ws.append(headers)

            # Formater les en-têtes
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Ajouter les données
            for item in history:
                ws.append([
                    item.get('numero_facture', ''),
                    str(item.get('date_vente', '')),
                    f"{item.get('montant_total', 0):.2f}",
                    item.get('statut', '')
                ])

            # Ajuster les largeurs
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15

            wb.save(filepath)
            return True, f"Export réussi : {filepath}"
        except Exception as e:
            return False, f"Erreur d'export : {str(e)}"


def export_sales_to_excel(sales, filepath):
    """Exporter les ventes en Excel"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventes"

        # En-têtes
        headers = ["N° Facture", "Client", "Date", "Montant TTC", "Montant Payé", "Montant Restant", "Statut"]
        ws.append(headers)

        # Formater les en-têtes
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajouter les données
        for sale in sales:
            ws.append([
                sale.get('numero_facture', ''),
                sale.get('client_nom', ''),
                str(sale.get('date_vente', ''))[:10],
                float(sale.get('montant_total', 0)),
                float(sale.get('montant_paye', 0)),
                float(sale.get('montant_reste', 0)),
                sale.get('statut', '')
            ])

        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12

        # Formater les montants
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=6):
            for cell in row:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right")

        # Colorer le statut
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
            for cell in row:
                if cell.value == 'payee':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell.value == 'partielle':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif cell.value == 'en_cours':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        wb.save(filepath)
        return True, f"Export réussi : {filepath}"
    except Exception as e:
        return False, f"Erreur d'export : {str(e)}"
